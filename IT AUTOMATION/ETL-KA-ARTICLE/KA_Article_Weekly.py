import os
from glob import glob
import pandas as pd
import os.path
import sys
from datetime import datetime,date, timedelta
import datetime
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook

PATH = sys.argv[1]
ArticleBaseOnRequest_Path = PATH +r'/ArticleBaseOnRequest.xlsx'
Person_Path = PATH +r'/Person.xlsx'
PersonGroup_Path = PATH + r'/PersonGroup.xlsx'
Article_Path = PATH +r'/Article.xlsx'
Request_Path = PATH +r'/Request.xlsx'
EXT = "*.zip"
all_zip_files = [file
                 for path, subdir, files in os.walk(PATH)
                 for file in glob(os.path.join(path, EXT))]
for x in all_zip_files:
    os.remove(x)
################################################ArticleBaseOnRequest_File################################################
EXT1 = "ArticleBaseOnRequest_*.csv"
all_ArticleBaseOnRequest_files = [file
                 for path, subdir, files in os.walk(PATH)
                 for file in glob(os.path.join(path, EXT1))]
df_from_ArticleBaseOnRequest_file = (pd.read_csv(f, sep=',') for f in all_ArticleBaseOnRequest_files)
df_ArticleBaseOnRequest_merged   = pd.concat(df_from_ArticleBaseOnRequest_file, ignore_index=True)
df_ArticleBaseOnRequest_merged.drop(df_ArticleBaseOnRequest_merged.columns[[0, 1]], axis = 1, inplace = True)
df=df_ArticleBaseOnRequest_merged.to_excel(ArticleBaseOnRequest_Path, index=False)
df=df_ArticleBaseOnRequest_merged.to_excel(ArticleBaseOnRequest_Path, index=False)
wb=load_workbook(ArticleBaseOnRequest_Path)
sheet=wb.active
for y in all_ArticleBaseOnRequest_files:
    os.remove(y)
print("ArticleBaseOnRequest is done")
################################################Person_ File ################################################
EXT2 = "Person_*.csv"
all_person_files = [file
                 for path, subdir, files in os.walk(PATH)
                 for file in glob(os.path.join(path, EXT2))]
df_from_person_file = (pd.read_csv(f, sep=',') for f in all_person_files)
df_person_merged   = pd.concat(df_from_person_file, ignore_index=True)
df_person_merged.drop(df_person_merged.columns[[0, 1]], axis = 1, inplace = True)
df=df_person_merged.to_excel(Person_Path, index=False)
wb=load_workbook(Person_Path)
sheet=wb.active
for y in all_person_files:
    os.remove(y)
print("Person file is done")
##############################################PersonGroup_ File ##############################################
EXT3 = "PersonGroup_*.csv"
all_persongroup_files = [file
                 for path, subdir, files in os.walk(PATH)
                 for file in glob(os.path.join(path, EXT3))]
df_from_persongroup_file = (pd.read_csv(f, sep=',') for f in all_persongroup_files)
df_persongroup_merged   = pd.concat(df_from_persongroup_file, ignore_index=True)
df_persongroup_merged.drop(df_persongroup_merged.columns[[0, 1]], axis = 1, inplace = True)
df=df_persongroup_merged.to_excel(PersonGroup_Path, index=False)
wb=load_workbook(PersonGroup_Path)
sheet=wb.active
#persongroup_max_row=sheet.max_row
#print(persongroup_max_row+1)
for z in all_persongroup_files:
    os.remove(z)
print("PersonGroup file is done")
################################################Article File ################################################
EXT4 = "Article_*.csv"
all_article_files = [file
                 for path, subdir, files in os.walk(PATH)
                 for file in glob(os.path.join(path, EXT4))]
df_from_article_file = (pd.read_csv(f, sep=',') for f in all_article_files)
df_article_merged   = pd.concat(df_from_article_file, ignore_index=True)
df_article_merged.drop(df_article_merged.columns[[0, 1]], axis = 1, inplace = True)
df=df_article_merged.to_excel(Article_Path, index=False)
wb=load_workbook(Article_Path)
sheet=wb.active
for y in all_article_files:
    os.remove(y)
print("Article file is done")
##############################################Request_ File ##############################################
EXT5 = "Request_*.csv"
all_Request_files = [file
                 for path, subdir, files in os.walk(PATH)
                 for file in glob(os.path.join(path, EXT5))]
df_from_Request_file = (pd.read_csv(f, sep=',') for f in all_Request_files)
df_Request_merged   = pd.concat(df_from_Request_file, ignore_index=True)
df_Request_merged.drop(df_Request_merged.columns[[0, 1]], axis = 1, inplace = True)
dict = {'Id':'Request ID','DisplayLabel':'Title','CurrentAssignment':'Current Assignment'}
df_Request_merged.rename(columns=dict,inplace=True)
df_Request_merged['Priority']=df_Request_merged['Priority'].replace('MediumPriority','Medium', regex=True)
df_Request_merged['Priority']=df_Request_merged['Priority'].replace('LowPriority','Low', regex=True)
df_Request_merged['Priority']=df_Request_merged['Priority'].replace('CriticalPriority','Critical', regex=True)
df_Request_merged['Priority']=df_Request_merged['Priority'].replace('HighPriority','High', regex=True)
df_Request_merged['ChatStatus']=df_Request_merged['ChatStatus'].replace('ChatStatusNone','None', regex=True)
df_Request_merged['ChatStatus']=df_Request_merged['ChatStatus'].replace('ChatStatusAbandoned','Abandoned', regex=True)
df_Request_merged['ChatStatus']=df_Request_merged['ChatStatus'].replace('ChatStatusPending','Pending', regex=True)
CreateTime=df_Request_merged['CreateTime']
CreateTime=pd.to_datetime(CreateTime, unit='ms')#.dt.strftime('%Y/%m/%d')
df_Request_merged['CreateTime'] = CreateTime
CloseTime=df_Request_merged['CloseTime']
CloseTime=pd.to_datetime(CloseTime, unit='ms')#.dt.strftime('%Y/%m/%d')
df_Request_merged['CloseTime'] = CloseTime
df=df_Request_merged.to_excel(Request_Path, index=False)
wb=load_workbook(Request_Path)
sheet=wb.active
for z in all_Request_files:
    os.remove(z)
print("Request file is done")
#############################################Vlookup for RequestedForPerson#################################################
df1 = pd.read_excel(Person_Path,engine='openpyxl')
dict = {'Id': 'RequestedForPerson','Name': 'Requested for Name'}
df1.rename(columns=dict,inplace=True)
df2=df1.append({'RequestedForPerson':'', 'Requested for Name':'#N/A '}, ignore_index=True)
df = df1.to_excel(Person_Path,engine='openpyxl',index=False)
df1 = pd.read_excel(Request_Path,engine='openpyxl')
df2 = pd.read_excel(Person_Path,engine='openpyxl')
join = pd.merge(df1,df2[['RequestedForPerson', 'Requested for Name']],on = 'RequestedForPerson' , how = 'left')
column = join.pop('RequestedForPerson')
df = join.to_excel(Request_Path,index=False)

print("vlookup1 is done")

#############################################Vlookup for AssignedToGroup#################################################
df1 = pd.read_excel(PersonGroup_Path,engine='openpyxl')
dict = {'Id': 'AssignedToGroup','Name': 'Assignment group Name'}
df1.rename(columns=dict,inplace=True)
df2=df1.append({'AssignedToGroup':'', 'Assignment group Name':'#N/A '}, ignore_index=True)
df = df1.to_excel(PersonGroup_Path,engine='openpyxl',index=False)
df1 = pd.read_excel(Request_Path,engine='openpyxl')
df2 = pd.read_excel(PersonGroup_Path,engine='openpyxl')
join = pd.merge(df1,df2[['AssignedToGroup', 'Assignment group Name']],on = 'AssignedToGroup' , how = 'left')
column = join.pop('AssignedToGroup')
df = join.to_excel(Request_Path,index=False)
print("vlookup2 is done")
#############################################Vlookup for AssignedToPerson#################################################
df1 = pd.read_excel(Person_Path,engine='openpyxl')
dict = {'RequestedForPerson': 'AssignedToPerson','Requested for Name': 'Assignee Name'}
df1.rename(columns=dict,inplace=True)
df2=df1.append({'AssignedToPerson':'', 'Assignee Name':'#N/A '}, ignore_index=True)
df = df1.to_excel(Person_Path,engine='openpyxl',index=False)
df1 = pd.read_excel(Request_Path,engine='openpyxl')
df2 = pd.read_excel(Person_Path,engine='openpyxl')
join = pd.merge(df1,df2[['AssignedToPerson', 'Assignee Name']],on = 'AssignedToPerson' , how = 'left')
column = join.pop('AssignedToPerson')
df = join.to_excel(Request_Path,index=False)
print("vlookup3 is done")
#############################################Vlookup for Knowledge Article ID#################################################
df1 = pd.read_excel(ArticleBaseOnRequest_Path,engine='openpyxl')
dict = {'firstEntity_Request': 'Request ID','secondEntity_Article': 'Knowledge Article ID'}
df1.rename(columns=dict,inplace=True)
df2=df1.append({'Request ID':'', 'Knowledge Article ID':'#N/A '}, ignore_index=True)
df = df1.to_excel(ArticleBaseOnRequest_Path,engine='openpyxl',index=False)
df1 = pd.read_excel(Request_Path,engine='openpyxl')
df2 = pd.read_excel(ArticleBaseOnRequest_Path,engine='openpyxl')
join = pd.merge(df1,df2[['Request ID', 'Knowledge Article ID']],on = 'Request ID' , how = 'left')
df = join.to_excel(Request_Path,index=False)
print("vlookup5 is done")
#############################################Vlookup for Knowledge Article NAME#################################################
df1 = pd.read_excel(Article_Path,engine='openpyxl')
dict = {'Id': 'Knowledge Article ID','Title': 'Knowledge Article Name'}
df1.rename(columns=dict,inplace=True)
df2=df1.append({'Knowledge Article ID':'', 'Knowledge Article Name':'#N/A '}, ignore_index=True)
df = df1.to_excel(Article_Path,engine='openpyxl',index=False)
df1 = pd.read_excel(Request_Path,engine='openpyxl')
df2 = pd.read_excel(Article_Path,engine='openpyxl')
join = pd.merge(df1,df2[['Knowledge Article ID', 'Knowledge Article Name']],on = 'Knowledge Article ID' , how = 'left')
df = join.to_excel(Request_Path,index=False)
print("vlookup6 is done")
######################################################Filter##############################################################################
# Get today's date
today = date.today().isoformat()
print("Today is: ", today)
  
# Yesterday date
yesterday = (date.today() - timedelta(1)).isoformat()
print("Yesterday was: ", yesterday)

#monday = (date.today() - timedelta(7)).isoformat() 
## just to run manually
monday = (date.today() - timedelta(11)).isoformat()
print("Last Monday's Date: ", monday)                 


start_date = monday
end_date = yesterday

df1 = pd.read_excel(Request_Path,engine='openpyxl')
mask = (df1['CreateTime'] > start_date) & (df1['CreateTime'] < today)
df = df1.loc[mask]
#print(df)
df2= df.to_excel(Request_Path,index=False)
print("filter is done")
##############################################Dropping extra col and keeping column in correct way###################################
df1 = pd.read_excel(Request_Path,engine='openpyxl')
df1.drop(df1.columns[[1,2,3,4,6,10,11]], axis = 1, inplace = True)
dict = {'CreateTime': 'Creation Time','CloseTime': 'Closed Time','Requested for Name':'Requested For','Assignment group Name':'Assignment Group'}
df1.rename(columns=dict,inplace=True)
df1=df1[["Request ID","Title","Priority","Requested For","Current Assignment","Assignment Group","Assignee Name","Knowledge Article ID","Knowledge Article Name","Creation Time","Closed Time"]]

dateFileName = monday = (date.today() - timedelta(7)).strftime("%d%B%Y")

finalRequestPath = PATH + "/Request_KA_Weekly_" + dateFileName + ".xlsx"
print(finalRequestPath)
df= df1.to_excel(finalRequestPath,index=False)
os.remove(Person_Path)
os.remove(PersonGroup_Path)
os.remove(ArticleBaseOnRequest_Path)
os.remove(Article_Path)
os.remove(Request_Path)
